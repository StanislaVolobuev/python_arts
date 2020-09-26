import os
import openpyxl


def open_weights():
    '''преобразует текстовый документ весов в словарь весов,
    возращая значения в виде словаря w_d'''
    weights = open(r'D:\data\weights.txt')
    w_d = []
    for row in weights:
        d = str(row)
        d = d.split()
        d.pop(1)
        d.pop(2)
        w_d.append(d)
    w_d = dict(w_d)
    return w_d


def open_text(d):
    '''принимает аргументом адрес текстового документа, возвращает
      количество строк в документе и значение'''
    n = 1
    session = open(d)

    for row in session:
        n += 1
        sq_v = [0, 0]
        if 20 < n:
            v = []

            m = list(row.split())
            v.append(float(m[1]))
            v.append(float(m[2]))
            sq_v[0] += abs(v[0] ** 2) / 2  # - mАргумент массы из файла весоы
            sq_v[1] += abs(v[1] ** 2) / 2

    return sq_v


def beat():
    d = open_weights()

    pach = r'D:\data'
    j = []
    for i in os.listdir(pach):  # где i так же имя испытуемого или иная папка
        if os.path.isdir(pach + '\\' + i):
            j.append(i)
    r_s = []  # Список участников с информацией о весе
    for i in j:
        if i in d.keys():
            print(i)
            r_s.append(i)
        else:
            print(i, '- нет информации о весе учачтника эксперемента')
    result = dict()
    for j in r_s:

        user = pach + '\\' + j  # Папака участника с сессиями

        for k in os.listdir(user):
            if k == 'vel':
                user_vel = user + '\\' + k
                result[j] = dict()
                for i in os.listdir(user_vel):
                    '''проработать и выделить номер'''

                    user_ses = user_vel + '\\' + i  # где i - хранит в себе имя участника и номер сессии участника
                    #n = i.split()
                    #n1 = n[1].split('.')
                    nom_ses = int(i[-5])
                    sq_v = list(open_text(user_ses))
                    res = []
                    for v in sq_v:
                        res.append(v * float(d[j]))
                        result[j][nom_ses] = res
    print(result)
    return result


print(beat())


def f_row():
    title = []
    x = '_X_'
    y = '_Y_'

    for nom in range(8):
        t = "En" + x + 'raw_00' + str(nom + 1)
        tt = "En" + y + 'raw_00' + str(nom + 1)
        if nom % 2 == 0:
            t = t + 'EC'
            tt = tt + 'EC'
        else:
            t = t + 'EO'
            tt = tt + 'EO'
        title.append(t)
        title.append(tt)
    return title


result = beat()
wb = openpyxl.Workbook()
wb.create_sheet(title='Первый лист', index=0)
sheet = wb['Первый лист']
t = f_row()
for col, name in zip(range(2, 20), t):
    cell = sheet.cell(row=1, column=col)
    cell.value = name
for row, word in zip(range(2, 20), result.keys()):
    cell = sheet.cell(row=row, column=1)
    cell.value = word
    for col, num in zip(range(2, 20, 2), range(1,9)):
        cell = sheet.cell(row=row, column=col)
        cell.value = result[word][num][0]
        col += 1
        cell = sheet.cell(row=row, column=col)
        cell.value = result[word][num][1]

wb.save('example1.xlsx')
