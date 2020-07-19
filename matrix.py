import openpyxl
import beat_folder
import first_row

result = beat_folder.beat()
# создаем новый excel-файл
wb = openpyxl.Workbook()

# добавляем новый лист
wb.create_sheet(title='Первый лист', index=0)
# получаем лист, с которым будем работать
sheet = wb['Первый лист']
t = first_row.f_row()
for col, name in zip(range(2, 20), t):
    cell = sheet.cell(row=1, column=col)
    cell.value = name
for row, word in zip(range(2, 20), result.keys()):
    cell = sheet.cell(row=row, column=1)
    cell.value = word
    for col, num in zip(range(2, 20, 2), result[word].keys()):
        cell = sheet.cell(row=row, column=col)
        cell.value = result[word][num][0][0]
        col += 1
        cell = sheet.cell(row=row, column=col)
        cell.value = result[word][num][0][1]

wb.save('example.xlsx')
