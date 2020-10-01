import os
import openpyxl


def open_weights():
    '''преобразует текстовый документ весов в словарь весов,
    возращая значения в виде словаря w_d'''
    weights = open(r'D:\data\weights.txt')
    weights_dict = dict()
    i = 0
    for row in weights:
        if i == 0:
            print(type(row))
        d = row.split()
        i += 1
        weights_dict[d[0]] = d[2]
    return weights_dict


def open_text(d):
    '''принимает аргументом адрес текстового документа, возвращает
      количество строк в документе и значение'''
    n = 0
    session = open(d)
    sq_v = [0, 0]
    m = [0, 0, 0]
    for row in session:
        n += 1


        if 19 < n:
            v = []
            m1 = list(row.split())
            v.append(float(m1[1]) - float(m[1]))
            v.append(float(m1[2]) - float(m[2]))
            m = m1
            sq_v[0] += ((v[0] ** 2) + (v[1] ** 2))/ 2  # - mАргумент массы из файла весоы
            # sq_v[1] += (v[1] ** 2) / 2

    return sq_v


def beat():
    weights_dict = open_weights()

    path = r'D:\data'
    j = []
    for i in os.listdir(path):  # где i так же имя испытуемого или иная папка
        if os.path.isdir(path + '\\' + i):
            j.append(i)
    r_s = []  # Список участников с информацией о весе
    for i in j:
        if i in weights_dict.keys():
            print(i)
            r_s.append(i)
        else:
            print(i, '- нет информации о весе учачтника эксперемента')
    result = dict()
    for j in r_s:

        user = path + '\\' + j  # Папака участника с сессиями

        for k in os.listdir(user):
            if k == 'vel':
                user_vel = user + '\\' + k
                result[j] = dict()
                for i in os.listdir(user_vel):

                    user_ses = user_vel + '\\' + i  # где i - хранит в себе имя участника и номер сессии участника
                    nom_ses = int(i[-5])
                    sq_v = list(open_text(user_ses))
                    res = []
                    for v in sq_v:
                        res.append(v * float(weights_dict[j]))
                        result[j][nom_ses] = res
    print(result)
    return result


result = beat()
num_sessions=result['Irina'].__len__()
print(num_sessions)
def head_row(num_sessions):  # (num_sessions):
    row_head = []
    # x = '_X_'
    # y = '_Y_'

    for num in range(num_sessions):  # range(num_sessions.__len__()):
        t = "En" +  'raw_00' + str(num + 1)
        # tt = "En" + y + 'raw_00' + str(num + 1)
        if num % 2 == 0:
            t = t + 'EC'
            # tt = tt + 'EC'
        else:
            t = t + 'EO'
            # tt = tt + 'EO'
        row_head.append(t)
        # row_head.append(tt)
    return row_head


wb = openpyxl.Workbook()
wb.create_sheet(title='Первый лист', index=0)
sheet = wb['Первый лист']
row_head = head_row(num_sessions)
for col, name in zip(range(2, num_sessions+2), row_head):
    cell = sheet.cell(row=1, column=col)
    cell.value = name
for row, key in zip(range(2, num_sessions+2), result.keys()):
    cell = sheet.cell(row=row, column=1)
    cell.value = key
    for col, num in zip(range(2, num_sessions+2), range(1, num_sessions+1)):  # num - номер сессии
        cell = sheet.cell(row=row, column=col)
        cell.value = result[key][num][0]
        # col += 1
        # cell = sheet.cell(row=row, column=col)
        # cell.value = result[key][num][1]

wb.save('energy_result.xlsx')

